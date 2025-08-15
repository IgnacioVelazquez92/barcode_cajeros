# app/controller.py
import subprocess
import sys
from tkinter import messagebox
from app import db
from app import pdf_generator
from app.utils import generar_clave, ofuscar_clave, normalizar_fecha
import os
from datetime import datetime
import sqlite3
from app.paths import base_dir, ensure_dirs
from openpyxl import load_workbook


def crear_cajero(
    nombre_completo: str,
    dni: str,
    legajo: str | None = None,
    nombre_sistema: str | None = None,
    sucursal: str | None = None,
    fecha_creacion: str | None = None,
) -> bool:
    """
    Alta con campos opcionales (legajo, nombre_sistema, sucursal, fecha_creacion).
    DNI debe ser único. La fecha se normaliza a YYYY-MM-DD (acepta DD/MM/YYYY).
    """
    if not nombre_completo or not dni or not dni.isdigit():
        messagebox.showerror(
            "Error", "Nombre completo y DNI numérico son requeridos.")
        return False

    nombre_completo = nombre_completo.strip().upper()
    legajo = legajo.strip().upper() if legajo else None
    nombre_sistema = nombre_sistema.strip().upper() if nombre_sistema else None
    sucursal = sucursal.strip().upper() if sucursal else None

    # Fecha
    try:
        fecha_norm = normalizar_fecha(fecha_creacion)
    except ValueError as ve:
        messagebox.showerror("Fecha inválida", str(ve))
        return False

    if db.existe_dni(dni):
        messagebox.showerror("Error", f"Ya existe un cajero con el DNI {dni}.")
        return False

    clave_base = generar_clave(dni)
    clave_guardar = ofuscar_clave(clave_base)

    try:
        db.insertar_cajero(
            nombre_completo, dni, clave_guardar,
            legajo=legajo, nombre_sistema=nombre_sistema, sucursal=sucursal,
            fecha_creacion=fecha_norm
        )
        return True
    except sqlite3.IntegrityError:
        messagebox.showerror("Error", f"Ya existe un cajero con el DNI {dni}.")
        return False
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo guardar el cajero: {e}")
        return False


def editar_cajero(
    id_cajero: int,
    nombre_completo: str,
    dni: str,
    legajo: str | None = None,
    nombre_sistema: str | None = None,
    sucursal: str | None = None,
    fecha_creacion: str | None = None,
) -> bool:
    """
    Edición completa, incluyendo fecha de creación.
    """
    if not nombre_completo or not dni.isdigit():
        messagebox.showerror(
            "Error", "Nombre completo y DNI válidos requeridos.")
        return False

    nombre_completo = nombre_completo.strip().upper()
    legajo = legajo.strip().upper() if legajo else None
    nombre_sistema = nombre_sistema.strip().upper() if nombre_sistema else None
    sucursal = sucursal.strip().upper() if sucursal else None

    try:
        fecha_norm = normalizar_fecha(
            fecha_creacion) if fecha_creacion is not None else None
    except ValueError as ve:
        messagebox.showerror("Fecha inválida", str(ve))
        return False

    clave_base = generar_clave(dni)
    clave_guardar = ofuscar_clave(clave_base)

    try:
        db.actualizar_cajero(
            id_cajero, nombre_completo, dni, clave_guardar,
            legajo=legajo, nombre_sistema=nombre_sistema, sucursal=sucursal,
            fecha_creacion=fecha_norm
        )
        return True
    except sqlite3.IntegrityError:
        messagebox.showerror(
            "Error", f"El DNI {dni} ya existe asignado a otro cajero.")
        return False
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo actualizar el cajero: {e}")
        return False


def actualizar_clave_personalizada(id_cajero: int, clave_plana: str, ofuscar: bool = True) -> bool:
    """
    Actualiza exclusivamente la clave del cajero.
    - Si ofuscar=True: aplica ofuscar_clave(clave_plana)
    - Si ofuscar=False: guarda tal cual (p.ej. 'pepe')
    Conserva legajo, nombre_sistema, sucursal y fecha_creacion sin cambios.
    """
    cajero = db.obtener_por_id(id_cajero)
    if not cajero:
        messagebox.showerror("Error", "Cajero no encontrado.")
        return False

    # Dataset: (id, legajo, nombre, nombre_sistema, dni, clave, fecha, sucursal)
    legajo = cajero[1]
    nombre = cajero[2]
    nombre_sistema = cajero[3]
    dni = cajero[4]
    sucursal = cajero[7]

    try:
        clave_a_guardar = ofuscar_clave(
            clave_plana) if ofuscar else clave_plana
        # No tocamos fecha_creacion (pasamos None para que COALESCE la mantenga)
        db.actualizar_cajero(
            id_cajero=id_cajero,
            nombre_completo=nombre,
            dni=dni,
            clave=clave_a_guardar,
            legajo=legajo,
            nombre_sistema=nombre_sistema,
            sucursal=sucursal,
            fecha_creacion=None,
        )
        return True
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo actualizar la clave: {e}")
        return False


def eliminar_cajero(id_cajero: int) -> bool:
    try:
        db.eliminar_cajero(id_cajero)
        return True
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo eliminar: {e}")
        return False


def buscar_cajeros(nombre: str):
    return db.buscar_por_nombre(nombre)


def obtener_cajeros():
    return db.obtener_todos()


def reimprimir_pdf(id_cajero: int) -> bool:
    cajero = db.obtener_por_id(id_cajero)
    if not cajero:
        messagebox.showerror("Error", "Cajero no encontrado.")
        return False

    # Dataset: (id, legajo, nombre, nombre_sistema, dni, clave, fecha, sucursal)
    nombre_completo = cajero[2] or ""
    nombre_sistema = cajero[3] or ""
    etiqueta = (nombre_sistema.strip() or nombre_completo.strip())
    clave = cajero[5]

    try:
        pdf_generator.generar_pdf(etiqueta, clave)
        messagebox.showinfo(
            "Reimpresión", f"PDF de {etiqueta} generado correctamente.")
        return True
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo generar el PDF: {e}")
        return False


def exportar_excel() -> bool:
    """
    Exporta todos los cajeros a un Excel en /excel con timestamp.
    - Si hay registros: los vuelca.
    - Si no hay registros: genera una PLANTILLA solo con encabezados (útil para importar luego).
    Abre la carpeta al finalizar. Si openpyxl no está disponible, hace fallback a PDF.
    Formato de columnas:
      ID | Legajo | Nombre completo | Nombre solutia | DNI | Clave | Fecha creación | Sucursal
    """
    from app.paths import base_dir, ensure_dirs
    ensure_dirs("excel")
    excel_dir = os.path.join(base_dir(), "excel")
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    ruta_xlsx = os.path.join(excel_dir, f"backup_{ts}.xlsx")

    # (id, legajo, nombre, nombre_sistema, dni, clave, fecha, sucursal)
    registros = db.obtener_todos()

    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Cajeros"

        # Encabezados (siempre)
        ws.append(["ID", "Legajo", "Nombre completo", "Nombre solutia",
                  "DNI", "Clave", "Fecha creación", "Sucursal"])

        # Filas (solo si hay)
        for r in registros:
            ws.append([
                r[0], r[1] or "", r[2] or "", r[3] or "", r[4] or "",
                r[5] or "", r[6] or "", r[7] or "",
            ])

        # Autoancho
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 42)

        wb.save(ruta_xlsx)
        if registros:
            messagebox.showinfo("Backup", f"Excel generado:\n{ruta_xlsx}")
        else:
            messagebox.showinfo(
                "Backup", f"No había registros.\nSe generó PLANTILLA Excel:\n{ruta_xlsx}")
        abrir_carpeta(excel_dir)
        return True

    except ImportError:
        # Fallback a PDF si falta openpyxl (incluye solo encabezados si no hay datos)
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import cm
        except Exception:
            messagebox.showerror(
                "Falta dependencia",
                "No se pudo generar Excel (openpyxl no instalado) ni PDF de respaldo.\n"
                "Instalá:\n  pip install openpyxl reportlab"
            )
            return False

        ruta_pdf = os.path.join(excel_dir, f"backup_{ts}.pdf")
        c = canvas.Canvas(ruta_pdf, pagesize=A4)
        page_w, page_h = A4
        left, top = 1.2 * cm, page_h - 1.2 * cm
        line_h = 0.55 * cm

        c.setFont("Helvetica-Bold", 10)
        c.drawString(
            left, top, "ID  |  Legajo  |  Nombre  |  Solutia  |  DNI  |  Clave  |  Fecha  |  Sucursal")
        c.setFont("Helvetica", 9)
        y = top - line_h

        if registros:
            for r in registros:
                texto = f"{r[0]} | {r[1] or ''} | {r[2] or ''} | {r[3] or ''} | {r[4] or ''} | {r[5] or ''} | {r[6] or ''} | {r[7] or ''}"
                c.drawString(left, y, texto[:180])
                y -= line_h

        c.showPage()
        c.save()
        if registros:
            messagebox.showinfo(
                "Backup", f"PDF de respaldo generado:\n{ruta_pdf}")
        else:
            messagebox.showinfo(
                "Backup", f"No había registros.\nSe generó PLANTILLA PDF:\n{ruta_pdf}")
        abrir_carpeta(excel_dir)
        return True

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo guardar el Excel:\n{e}")
        return False


def abrir_carpeta(path: str) -> bool:
    try:
        if sys.platform.startswith("win"):
            os.startfile(path)  # type: ignore
        elif sys.platform == "darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])
        return True
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo abrir la carpeta:\n{e}")
        return False


def abrir_carpeta_pdfs() -> bool:
    from app.pdf_generator import PDF_DIR
    return abrir_carpeta(PDF_DIR)


def importar_desde_excel(ruta_xlsx: str) -> dict:
    """
    Importa un Excel con el MISMO formato que el backup exportado:
    Encabezados esperados (case-insensitive):
      ID | Legajo | Nombre completo | Nombre solutia | DNI | Clave | Fecha creación | Sucursal
    - Si 'Clave' viene vacía: se genera automáticamente (DNI+fecha -> ofuscada).
    - Si 'Clave' viene con valor: se guarda tal cual (sin re-ofuscar).
    - DNI duplicado: se omite y se registra en el log.
    Devuelve: {'insertados': int, 'saltados': int, 'log': ruta_log}
    """
    wb = load_workbook(ruta_xlsx, data_only=True)
    ws = wb.active

    # Mapear encabezados (case-insensitive)
    header_row = [(str(c.value).strip() if c.value is not None else "")
                  for c in ws[1]]
    hdr_idx = {h.upper(): i for i, h in enumerate(header_row)}

    def col(*names):
        for n in names:
            u = n.upper()
            if u in hdr_idx:
                return hdr_idx[u]
        return None

    idx_id = col("ID")
    idx_legajo = col("LEGAJO")
    idx_nombre = col("NOMBRE COMPLETO", "NOMBRE")
    idx_solutia = col("NOMBRE SOLUTIA", "NOMBRE SISTEMA")
    idx_dni = col("DNI")
    idx_clave = col("CLAVE")
    idx_fecha = col("FECHA CREACIÓN", "FECHA CREACION", "FECHA")
    idx_sucursal = col("SUCURSAL")

    requeridos = [idx_nombre, idx_dni]
    if any(i is None for i in requeridos):
        messagebox.showerror(
            "Importar Excel", "Encabezados mínimos no encontrados: 'Nombre completo' y 'DNI'.")
        return {"insertados": 0, "saltados": 0, "log": ""}

    ensure_dirs("excel")
    log_dir = os.path.join(base_dir(), "excel")
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    ruta_log = os.path.join(log_dir, f"import_log_{ts}.txt")

    insertados = 0
    saltados = 0
    logs = []

    # Iterar filas
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v in (None, "") for v in row):
            continue

        def val(idx):
            if idx is None:
                return None
            v = row[idx]
            return "" if v is None else str(v).strip()

        nombre = (val(idx_nombre) or "").upper()
        dni = (val(idx_dni) or "")
        legajo = (val(idx_legajo) or None)
        if legajo:
            legajo = legajo.upper()
        solutia = (val(idx_solutia) or None)
        if solutia:
            solutia = solutia.upper()
        sucursal = (val(idx_sucursal) or None)
        if sucursal:
            sucursal = sucursal.upper()

        if not nombre:
            saltados += 1
            logs.append("Fila omitida: NOMBRE vacío.")
            continue
        if not dni.isdigit():
            saltados += 1
            logs.append(f"Fila omitida: DNI inválido ({dni}).")
            continue
        if db.existe_dni(dni):
            saltados += 1
            logs.append(f"Fila omitida: DNI duplicado en base ({dni}).")
            continue

        # Fecha
        raw_fecha = row[idx_fecha] if idx_fecha is not None else None
        if isinstance(raw_fecha, datetime):
            fecha_norm = raw_fecha.strftime("%Y-%m-%d")
        else:
            try:
                fecha_norm = normalizar_fecha(
                    str(raw_fecha) if raw_fecha not in (None, "") else None)
            except ValueError as ve:
                saltados += 1
                logs.append(
                    f"Fila omitida: fecha inválida ({raw_fecha}). {ve}")
                continue

        # Clave
        raw_clave = val(idx_clave) if idx_clave is not None else ""
        if raw_clave:
            clave = raw_clave  # guardar tal cual
        else:
            # generar a partir de DNI (comportamiento estándar)
            clave = ofuscar_clave(generar_clave(dni))

        try:
            db.insertar_cajero(
                nombre_completo=nombre,
                dni=dni,
                clave=clave,
                legajo=legajo,
                nombre_sistema=solutia,
                sucursal=sucursal,
                fecha_creacion=fecha_norm
            )
            insertados += 1
        except Exception as e:
            saltados += 1
            logs.append(f"Error al insertar DNI {dni}: {e}")

    # Guardar log
    with open(ruta_log, "w", encoding="utf-8") as f:
        f.write("== Importación desde Excel ==\n")
        f.write(f"Archivo: {ruta_xlsx}\n")
        f.write(f"Insertados: {insertados} | Saltados: {saltados}\n\n")
        for ln in logs:
            f.write(ln + "\n")

    messagebox.showinfo("Importación",
                        f"Insertados: {insertados}\nSaltados: {saltados}\n\nLog:\n{ruta_log}")
    abrir_carpeta(log_dir)
    return {"insertados": insertados, "saltados": saltados, "log": ruta_log}
